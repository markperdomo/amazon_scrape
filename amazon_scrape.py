#!env/bin/python

"""
AmazonScrape

Usage: amazon_scrape.py [--number=<count>] [--output=output.xlsx]

Select and scrape a given number of Amazon book listings.
"""

from docopt import docopt
from api_key import *
import requests
from bs4 import BeautifulSoup
import json
import time
import bottlenose
from openpyxl import Workbook


        
def random_amazon_link(number):
    """Use the site http://www.bookbookgoose.com/ to generate a random
    amazon link.  Grab the page and slice it apart for a link"""
    r = requests.get(
        'http://www.bookbookgoose.com/v1/get_books?n={0}'.format(number)
        )
    response = r.json()
    books = []
    for item in response:
        book = {
        'author':item[0],
        'title':item[1],
        'link':item[2],
        }
        books.append(book)
    return books
        
def grab_amazon_data(link):
    """Grab data from an amazon link.  Data to get:
        1) Price (amazon price-watch for hardback)
        2) Product score/user rating
        3) Number of ratings
        4) Score of first user rating
        5) Date of first user rating
        6) Shipping weight
        7) Time since first user rating
        8) Prime Product or not

    return as dictionary of values
    """
    amazon = bottlenose.Amazon(
        AWS_ACCESS_KEY_ID, 
        AWS_SECRET_ACCESS_KEY, 
        AWS_ASSOCIATE_TAG,
        )
    product = {}
    response = amazon.ItemLookup(ItemId=find_amazon_id(link),  ResponseGroup="Large")
    soup = BeautifulSoup(response, "html.parser")
    product['sales_rank'] = int(soup.salesrank.string)
    product['title'] = soup.title.string
    product['author'] = soup.author.string
    product['binding'] = soup.binding.string
    product['has_reviews'] = soup.hasreviews.string
    product['reviews_url'] = soup.customerreviews.iframeurl.string
    product['ship_weight'] = float(soup.find(units="hundredths-pounds").string)/100.00
    product['price'] = soup.listprice.formattedprice.string
    product['product_url'] = link

    return product


def find_amazon_id(link):
    """Find amazon item id from a passed link
    ONLY WORKS FOR BOOKS RIGHT NOW
    sample book url:
    http://www.amazon.com/Carbon-isotope-fractionation-trophic-transfer/dp/B000RR3CXS%3FSubscriptionId%3D1XJTRNMGKSD3T57YM002%26tag%3Dquasika-20%26linkCode%3Dxm2%26camp%3D2025%26creative%3D165953%26creativeASIN%3DB000RR3CXS
    """
    return link.split('/dp/')[1].split('%3F')[0]

def build_excel_doc(data,filename="output.xlsx"):
    """Take the data and output an excel file"""
    print 'Building workbook....',
    wb = Workbook(guess_types=True)
    ws = wb.active
    ws.title = 'Products'
    #fill in column headers
    for c,k in enumerate(data[0]):
        col = c + 1
        _ = ws.cell(column=col, row=1, value=k)
        #for each column fill in data down the line
        for r,v in enumerate(data):
            row = r + 2
            _ = ws.cell(column=col, row=row, value=v[k])
            


    wb.save(filename)
    print 'done!'

def main():
    arguments = docopt(__doc__)
    count = arguments['--number'] or 10
    print 'Will grab ' + str(count) + ' links from Amazon'
    output_data = []
    while len(output_data) < count:
        links = random_amazon_link(max(50,count))
        for amazon_link in links:
            if len(output_data) < count:
                try:
                    data = grab_amazon_data(amazon_link['link'])
                    output_data.append(data)
                    print 'Successfully grabbed ' + data['title']
                except:
                    print 'ERROR GRABBING #' + find_amazon_id(amazon_link['link'])
                time.sleep(1.01)

    build_excel_doc(output_data,filename="output.xlsx")


if __name__ == '__main__':
    main()